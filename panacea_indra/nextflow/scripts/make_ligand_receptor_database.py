import os
import sys
import tqdm
import pickle
import logging
import datetime
import openpyxl
import pandas as pd
from indra.util import batch_iter
from collections import defaultdict
from indra.statements import Complex
from indra.sources import indra_db_rest
import indra.tools.assemble_corpus as ac
from indra.ontology.bio import bio_ontology
from indra.databases.uniprot_client import um
from indra.assemblers.html import HtmlAssembler
# from make_enzyme_receptor_db import get_all_enzymes
from indra.sources.omnipath import process_from_web
from indra.assemblers.cx.assembler import CxAssembler
from indra.databases import uniprot_client, hgnc_client
from indra_db.client.principal.curation import get_curations
from indra.databases.hgnc_client import get_hgnc_from_mouse, get_hgnc_name

logger = logging.getLogger('receptor_ligand_interactions')

mouse_gene_name_to_mgi = {v: um.uniprot_mgi.get(k)
                          for k, v in um.uniprot_gene_name.items()
                          if k in um.uniprot_mgi}

up_hgnc = {v: k for k, v in um.uniprot_gene_name.items()
           if k in um.uniprot_hgnc}

db_curations = get_curations()

#__file__ = '/Users/sbunga/gitHub/panacea_indra/panacea_indra/nextflow/scripts/make_ligand_receptor_database.py'
HERE = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(HERE, os.pardir, 'input')
OUTPUT = os.path.join(HERE, os.pardir, 'output')
INDRA_DB_PKL = os.path.join(INPUT, 'db_dump_df.pkl')
NATURE_LIGAND_RECEPTOR_SPREADSHEET = os.path.join(INPUT, 'ncomms8866-s3.xlsx')
GO_ANNOTATIONS = os.path.join(INPUT, 'goa_human.gaf')
ION_CHANNELS = os.path.join(INPUT, 'ion_channels.txt')
SURFACE_PROTEINS_WB = os.path.join(INPUT, 'Surface Proteins.xlsx')


def _load_goa_gaf():
    """Load the gene/GO annotations as a pandas data frame."""
    # goa_ec = {'EXP', 'IDA', 'IPI', 'IMP', 'IGI', 'IEP', 'HTP', 'HDA', 'HMP',
    #          'HGI', 'HEP', 'IBA', 'IBD'}
    goa = pd.read_csv(GO_ANNOTATIONS, sep='\t',
                      skiprows=23, dtype=str,
                      header=None,
                      names=['DB',
                             'DB_ID',
                             'DB_Symbol',
                             'Qualifier',
                             'GO_ID',
                             'DB_Reference',
                             'Evidence_Code',
                             'With_From',
                             'Aspect',
                             'DB_Object_Name',
                             'DB_Object_Synonym',
                             'DB_Object_Type',
                             'Taxon',
                             'Date',
                             'Assigned',
                             'Annotation_Extension',
                             'Gene_Product_Form_ID'])
    goa = goa.sort_values(by=['DB_ID', 'GO_ID'])
    # Filter out all "NOT" negative evidences
    goa['Qualifier'].fillna('', inplace=True)
    goa = goa[~goa['Qualifier'].str.startswith('NOT')]
    # Filter to rows with evidence code corresponding to experimental
    # evidence
    # goa = goa[goa['Evidence_Code'].isin(goa_ec)]
    return goa


goa = _load_goa_gaf()


def get_pain_mol():
    PAIN_SIGNAL_MOL = {
        "Prostaglandins": "CHEBI:26333",
        "Brandykinin": "CHEBI:3165"
    }

    CHEBI_LIST = {}
    CHEBI_NAMES = {}
    for compounds, chebi_id in PAIN_SIGNAL_MOL.items():
        CHEBI_LIST[compounds] = \
            [children[1] for children in
             bio_ontology.get_children('CHEBI',
                                       chebi_id)]

        CHEBI_NAMES[compounds] = \
            [bio_ontology.get_name('CHEBI', ids)
             for ids in CHEBI_LIST[compounds]]

    return CHEBI_NAMES


PAIN_MOL_NAMES = get_pain_mol()


def load_indra_df(fname):
    """Return an INDRA Statement data frame from a pickle file."""
    logger.info('Loading INDRA DB dataframe')
    with open(fname, 'rb') as fh:
        df = pickle.load(fh)
    logger.info('Loaded %d rows from %s' % (len(df), fname))
    return df


# Load the INDRA DB DF
indra_df = load_indra_df(INDRA_DB_PKL)


def get_hashes_by_gene_pair(df, ligand_genes, receptor_genes):
    hashes_by_gene_pair = defaultdict(set)

    for a, b, hs in zip(df.agA_name, df.agB_name, df.stmt_hash):
        if a in ligand_genes and b in receptor_genes:
            hashes_by_gene_pair[(a, b)].add(hs)
    return hashes_by_gene_pair


def download_statements(hashes):
    """Download the INDRA Statements corresponding to a set of hashes.
    """
    stmts_by_hash = {}
    for group in tqdm.tqdm(batch_iter(hashes, 200), total=int(len(hashes) / 200)):
        idbp = indra_db_rest.get_statements_by_hash(list(group),
                                                    ev_limit=10)
        for stmt in idbp.statements:
            stmts_by_hash[stmt.get_hash()] = stmt
    return stmts_by_hash


def get_genes_for_go_ids(go_ids):
    """Return genes that are annotated with a given go ID or its children."""
    df = goa[goa['GO_ID'].isin(set(go_ids))]
    up_ids = sorted(list(set(df['DB_ID'])))
    gene_names = [uniprot_client.get_gene_name(up_id) for up_id in up_ids]
    gene_names = {g for g in gene_names if g}
    return gene_names


def fix_dates(gene_names):
    replacements = {
        datetime.datetime(2020, 3, 7, 0, 0): 'March7',
        datetime.datetime(2020, 3, 2, 0, 0): 'March2',
        datetime.datetime(2020, 3, 4, 0, 0): 'March4',
        datetime.datetime(2020, 3, 5, 0, 0): 'March5',
        datetime.datetime(2020, 3, 6, 0, 0): 'March6',
        datetime.datetime(2020, 3, 9, 0, 0): 'March9',
        datetime.datetime(2020, 3, 8, 0, 0): 'March8',
        datetime.datetime(2020, 3, 11, 0, 0): 'Mar11',
        datetime.datetime(2020, 9, 1, 0, 0): 'Sept1',
        datetime.datetime(2020, 9, 2, 0, 0): 'Sept2',
        datetime.datetime(2020, 9, 3, 0, 0): 'Sept3',
        datetime.datetime(2020, 9, 4, 0, 0): 'Sept4',
        datetime.datetime(2020, 9, 5, 0, 0): 'Sept5',
        datetime.datetime(2020, 9, 6, 0, 0): 'Sept6',
        datetime.datetime(2020, 9, 7, 0, 0): 'Sept7',
        datetime.datetime(2020, 9, 8, 0, 0): 'Sept8',
        datetime.datetime(2020, 9, 9, 0, 0): 'Sept9',
        datetime.datetime(2020, 9, 10, 0, 0): 'Sept10',
        datetime.datetime(2020, 9, 11, 0, 0): 'Sept11',
        datetime.datetime(2020, 9, 15, 0, 0): 'Sept15',
    }
    fixed_gene_names = set()
    for gene_name in gene_names:
        if isinstance(gene_name, datetime.datetime):
            fixed_gene_names.add(replacements[gene_name])
        else:
            fixed_gene_names.add(gene_name)
    return fixed_gene_names


def read_workbook(workbook):
    """ This function takes Excel workbook as an input and
    returns ligand and receptor gene list respectively.
    Input: Excel workbook with single(2 columns) or two sheets
    Condition: considers first column/sheet as ligand genes and second
    column/shet as receptor genes
    """
    ligands_sheet = 'updated list of ligands '
    receptors_sheet = 'RPKM > 1.5 cfiber'
    wb = openpyxl.load_workbook(workbook)
    ligands = fix_dates(set([row[0].value for row in wb[ligands_sheet]][1:]))
    receptors = fix_dates(set([row[0].value
                               for row in wb[receptors_sheet]][1:]))
    return ligands, receptors


def read_gene_list(infile, mode):
    gene_list = []
    try:
        with open(infile, mode) as FH:
            for eachGene in FH:
                gene_list.append(eachGene.strip("\n"))
        return gene_list

    except FileNotFoundError:
        sys.exit("Given file doesn't exist")


def filter_complex_statements(stmts, ligands, receptors):
    filtered_stmts = []
    for stmt in stmts:
        if isinstance(stmt, Complex):
            if len(stmt.members) <= 2:
                if (any(a.name in ligands for a in stmt.members)
                        and any(a.name in receptors for a in stmt.members)):
                    filtered_stmts.append(stmt)
        else:
            filtered_stmts.append(stmt)

    return filtered_stmts


def filter_op_stmts(op_stmts, lg, rg):
    """ Filter out the statements which are not ligand and receptor """
    logger.info(f'Filtering {len(op_stmts)} to ligand-receptor interactions')
    filtered_stmts = [stmt for stmt in op_stmts if
                      (any(a.name in lg for a in stmt.agent_list())
                       and any(a.name in rg for a in stmt.agent_list()))]
    logger.info(f'{len(filtered_stmts)} left after filter')
    return filtered_stmts


def html_assembler(indra_stmts, fname):
    """Assemble INDRA statements into a HTML report"""
    html_assembler = HtmlAssembler(indra_stmts,
                                   db_rest_url='https://db.indra.bio')
    assembled_html_report = html_assembler.make_model(no_redundancy=True)
    html_assembler.save_model(fname)
    return assembled_html_report


def cx_assembler(indra_stmts, fname):
    """Assemble INDRA statements into a CX report"""
    cx_assembler = CxAssembler(indra_stmts)
    assembled_cx_report = cx_assembler.make_model()
    cx_assembler.save_model(fname)
    ndex_network_id = cx_assembler.upload_model(ndex_cred=None,
                                                private=True, style='default')
    return assembled_cx_report, ndex_network_id


def get_receptor_by_ligands(receptors_in_data, ligands_in_data, stmts):
    receptor_by_ligands = defaultdict(set)
    for stmt in stmts:
        agent_names = {agent.name for agent in stmt.agent_list()}
        receptors = agent_names & receptors_in_data
        ligands = agent_names & ligands_in_data
        for receptor in receptors:
            receptor_by_ligands[receptor] |= ligands
    return dict(receptor_by_ligands)


def filter_out_medscan(stmts):
    logger.info('Filtering out medscan evidence on %d statements' % len(stmts))
    new_stmts = []
    for stmt in stmts:
        new_evidence = [e for e in stmt.evidence if e.source_api != 'medscan']
        if not new_evidence:
            continue
        stmt.evidence = new_evidence
        if not stmt.evidence:
            continue
        new_stmts.append(stmt)
    logger.info('%d statements after filter' % len(new_stmts))
    return new_stmts


def filter_db_only(stmts):
    new_stmts = []
    for stmt in stmts:
        sources = {ev.source_api for ev in stmt.evidence}
        if sources <= {'reach', 'sparser', 'trips', 'rlimsp', 'medscan', 'eidos'}:
            continue
        new_stmts.append(stmt)
    return new_stmts


def process_seurat_csv(infile, fc):
    """ Process Seurat dataframe and only filter in
    genes with the given Fold change """
    l_df = pd.read_csv(infile, header=0, sep=",")
    l_df.columns = l_df.columns.str.replace('Unnamed: 0', 'Genes')
    filtered_df = l_df[l_df['avg_logFC'] > 0.25][['Genes', 'avg_logFC']]
    filtered_df = filtered_df.sort_values(by='avg_logFC', ascending=False)
    filtered_dict = {}
    for r, c in filtered_df.iterrows():
        filtered_dict[c[1]] = c[0]
    # return set(filtered_markers)
    return filtered_dict


def get_de_product_list(de_enzyme_product_list,
                        de_enzyme_stmts):
    if len(de_enzyme_product_list) > 1:
        de_enzyme_product_list = pd.merge(de_enzyme_stmts, de_enzyme_product_list,
                                          on=['Enzyme', 'Interaction', 'product', 'logFC'],
                                          how="outer").fillna('')
        return de_enzyme_product_list.sort_values(by='logFC', ascending=False)

    elif len(de_enzyme_product_list) < 1:
        de_enzyme_product_list = de_enzyme_stmts
        return de_enzyme_product_list


def filter_incorrect_curations(stmts):
    # Filter incorrect curations
    indra_op_filtered = ac.filter_by_curation(stmts,
                                              curations=db_curations)
    return indra_op_filtered


def ligand_mgi_to_hgnc_name(seurat_ligand_genes):
    filtered_mgi = defaultdict(set)
    for logfc, gene in seurat_ligand_genes.items():
        if gene in mouse_gene_name_to_mgi:
            filtered_mgi[(gene, logfc)].add(mouse_gene_name_to_mgi[gene])

    hgnc_gene_dict = defaultdict(set)
    seen_genes = set()
    for key, value in filtered_mgi.items():
        mgi_id = next(iter(value))
        hgnc_id = get_hgnc_from_mouse(mgi_id)
        hgnc_symbol = get_hgnc_name(hgnc_id)
        if hgnc_symbol not in seen_genes:
            hgnc_gene_dict[(key[1])].add(hgnc_symbol)
        else:
            pass
        seen_genes.add(hgnc_symbol)
    return hgnc_gene_dict


def mgi_to_hgnc_name(gene_list):
    """Convert given mouse gene symbols to HGNC equivalent symbols"""
    filtered_mgi = {mouse_gene_name_to_mgi[gene] for gene in gene_list
                    if gene in mouse_gene_name_to_mgi}
    hgnc_gene_set = set()
    for mgi_id in filtered_mgi:
        hgnc_id = get_hgnc_from_mouse(mgi_id)
        hgnc_gene_set.add(get_hgnc_name(hgnc_id))
    return hgnc_gene_set


def process_df(workbook):
    wb = openpyxl.load_workbook(workbook)
    df = {
        'ligands': [row[1].value for row in wb['All.Pairs']][1:],
        'receptors': [row[3].value for row in wb['All.Pairs']][1:]
    }
    lg_rg = pd.DataFrame(df)
    return lg_rg


def expand_with_child_go_terms(terms):
    all_terms = set(terms)
    for term in terms:
        child_terms = bio_ontology.get_children('GO', term)
        all_terms |= {c[1] for c in child_terms}
    return all_terms


def get_receptors():
    receptor_terms = ['signaling receptor activity']
    receptor_go_ids = [bio_ontology.get_id_from_name('GO', term)[1]
                       for term in receptor_terms]
    receptor_go_ids = expand_with_child_go_terms(receptor_go_ids)
    # Filtering out the nuclear receptors from the receptor list
    receptor_go_ids = {r for r in receptor_go_ids if 'receptor' in
                       bio_ontology.get_name('GO', r) or 'sensor' in
                       bio_ontology.get_name('GO', r) or 'channel' in
                       bio_ontology.get_name('GO', r)} - \
                      expand_with_child_go_terms(['GO:0004879'])
    receptor_genes_go = get_genes_for_go_ids(receptor_go_ids)
    receptor_genes_go -= {'NR2C2', 'EGF'}
    # Add ION channels to the receptor list
    ion_channels = set()
    with open(ION_CHANNELS, 'r') as fh:
        for line in fh:
            ion_channels.add(line.strip())
    receptor_genes_go |= ion_channels
    return receptor_genes_go


def get_ligands():
    # Read and extract cell surface proteins from CSPA DB
    wb = openpyxl.load_workbook(SURFACE_PROTEINS_WB)
    surface_protein_set = set(row[4].value for row in wb['Sheet 1']
                              if row[6].value == 'yes')
    logger.info('Got %d surface proteins from spreadsheet' %
                len(surface_protein_set))
    ligand_terms = ['cytokine activity', 'hormone activity',
                    'growth factor activity',
                    'extracellular matrix structural constituent']
    # Getting GO id's for ligands and receptors by using
    # GO terms
    ligand_go_ids = [bio_ontology.get_id_from_name('GO', term)[1]
                     for term in ligand_terms]
    ligand_go_ids = expand_with_child_go_terms(ligand_go_ids)

    # Converting GO id's to gene symbols
    ligand_genes_go = get_genes_for_go_ids(ligand_go_ids)
    manual_ligands = set()
    return surface_protein_set | ligand_genes_go | manual_ligands


def make_interaction_df(indra_op):
    # Make INDRA_DB interactions dataframe
    interactions = []
    for receptor, ligands in indra_op.items():
        for ligand in ligands:
            interactions.append(
                {
                    'ligands': ligand,
                    'receptors': receptor,
                    'interactions': ligand + '_' + receptor
                }
            )
    interaction_df = pd.DataFrame(interactions)
    return interaction_df


def _make_nature_df(nature_interactions):
    # Make nature interactions dataframe
    nature_interactions_df = []
    for rows, cols in nature_interactions.iterrows():
        nature_interactions_df.append(
            {
                'ligands': cols[0],
                'receptors': cols[1],
                'interactions': cols[0] + '_' + cols[1]

            }
        )
    nature_interactions_df = pd.DataFrame(nature_interactions_df)
    return nature_interactions_df


def process_nature_paper():
    nature_xlsx = process_df(os.path.join(INPUT, 'ncomms8866-s3.xlsx'))
    nature_lg_by_rg = defaultdict(set)

    nature_dataframe = []
    count = 0

    for rows, cols in nature_xlsx.iterrows():
        nature_lg_by_rg[(cols[0])].add(cols[1])
        count += 1
        if cols[0] in up_hgnc and cols[1] in up_hgnc:
            nature_dataframe.append(
                {
                    'id_cp_interaction': 'NATURE-' + str(count),
                    'partner_a': up_hgnc[cols[0]],
                    'partner_b': up_hgnc[cols[1]],
                    'source': 'NATURE'
                }
            )

    nature_dataframe = pd.DataFrame(nature_dataframe)

    nature_dataframe.to_csv(os.path.join(HERE, os.pardir, 'output', 'nature_uniprot.csv'),
                            sep=",", index=False)
    return _make_nature_df(nature_xlsx)


def merge_interactions(interactions, genes_file, uniprot_file):
    df = [{'partner_a': i.split('_')[0],
           'partner_b': i.split('_')[1]}
          for i in interactions]

    interactions_hgnc = pd.DataFrame(df)
    interactions_hgnc.to_csv(os.path.join(HERE, 'output', genes_file),
                             sep=",", index=False)

    cellphonedb_df = []
    count = 0
    for r, c in interactions_hgnc.iterrows():
        count += 1
        if c[0] in up_hgnc and c[1] in up_hgnc:
            cellphonedb_df.append(
                {
                    'id_cp_interaction': 'Woolf-' + str(count),
                    'partner_a': up_hgnc[c[0]],
                    'partner_b': up_hgnc[c[1]],
                    'source': 'INDRA'
                }
            )

    cellphonedb_df = pd.DataFrame(cellphonedb_df)
    cellphonedb_df.to_csv(os.path.join(HERE, 'output', uniprot_file),
                          sep=",", index=False)


def filter_to_complex_statements(stmts, ligands, receptors):
    ''' Filter statements to only complex type '''

    readers = {'medscan', 'eidos', 'reach',
               'rlimsp', 'trips', 'sparser',
               'tees', 'geneways', 'isi'}

    filtered_stmts = []
    for stmt in stmts:
        if filter_to_bel(stmt):
            filtered_stmts.append(stmt)
            continue
        if isinstance(stmt, Complex):
            if len(stmt.members) <= 2:
                if (any(a.name in ligands for a in stmt.members)
                        and any(a.name in receptors for a in stmt.members)):

                    sources = {ev.source_api for ev in stmt.evidence}
                    evidence = len(stmt.evidence)

                    if len(sources) == 1 and 'sparser' in sources:
                        continue
                    if evidence < 2 and sources <= readers:
                        continue
                    filtered_stmts.append(stmt)
    return filtered_stmts


def filter_to_bel(stmt):
    sources = {ev.source_api for ev in stmt.evidence}
    if ('bel' in sources) and (type(stmt).__name__ == 'IncreaseAmount' or
                               type(stmt).__name__ == 'Activation'):
        for ev in stmt.evidence:
            if 'bel' in ev.source_api and ev.epistemics.get('direct') == True:
                return stmt


if __name__ == '__main__':
    receptor_genes_go = get_receptors()
    # remove all the receptors from the surface_protein_set
    full_ligand_set = get_ligands() - receptor_genes_go

    if not os.path.isfile(os.path.join(OUTPUT, 'indra_op_stmts.pkl')):
        # Now get INDRA DB Statements for the receptor-ligand pairs
        hashes_by_gene_pair = get_hashes_by_gene_pair(indra_df, full_ligand_set,
                                                      receptor_genes_go)

        # get the union of all the statement hashes
        all_hashes = set.union(*hashes_by_gene_pair.values())
        # Download the statements by hashes
        stmts_by_hash = download_statements(all_hashes)
        # get only the list of all the available statements
        indra_db_stmts = list(stmts_by_hash.values())

        # Filtering out the indirect INDRA statements
        indra_db_stmts = ac.filter_direct(indra_db_stmts)
        logger.info('Total statements after filtering to direct ones %d' % (len(indra_db_stmts)))

        # Filter out the statements to database only
        # indra_db_stmts = filter_db_only(indra_db_stmts)
        # logger.info('Total statements after filtering to database only %d' % (len(indra_db_stmts)))

        # Fetch omnipath database biomolecular interactions and
        # process them into INDRA statements
        op = process_from_web()
        logger.info('Total OP statements %d' % (len(op.statements)))

        # Filter statements which are not ligands/receptors from
        # OmniPath database
        op_filtered = filter_op_stmts(op.statements, full_ligand_set,
                                      receptor_genes_go)
        op_filtered = ac.filter_direct(op_filtered)

        op_filtered = ac.filter_by_curation(op_filtered,
                                            curations=db_curations)

        # Merge omnipath/INDRA statements and run assembly
        indra_op_stmts = ac.run_preassembly(indra_db_stmts + op_filtered,
                                            run_refinement=False)
        with open(os.path.join(OUTPUT, 'indra_op_stmts.pkl'), 'wb') as fh:
            pickle.dump(indra_op_stmts, fh)
    else:
        with open(os.path.join(OUTPUT, 'indra_op_stmts.pkl'), 'rb') as fh:
            indra_op_stmts = pickle.load(fh)

    # Filter incorrect curations        
    indra_op_filtered = filter_incorrect_curations(indra_op_stmts)

    # Filter to complex statements
    indra_op_filtered = filter_to_complex_statements(indra_op_filtered,
                                                     full_ligand_set,
                                                     receptor_genes_go)

    # We do this again because when removing complex members, we
    # end up with more duplicates
    indra_op_filtered = ac.run_preassembly(indra_op_filtered,
                                           run_refinement=False)

    # get ligands by receptor for indra_op
    indra_op_receptor_by_ligands = get_receptor_by_ligands(receptor_genes_go,
                                                           full_ligand_set,
                                                           indra_op_filtered)

    # Assemble the statements into HTML formatted report and save into a file
    indra_db_html_report = \
        html_assembler(indra_op_filtered,
                       fname=(os.path.join(HERE, os.pardir, 'output', 'indra_op_interactions.html')))

    nature_interactions = process_nature_paper()
    indra_db_interactions = make_interaction_df(indra_op_receptor_by_ligands)

    # Create a cellphone db formatted indra_op database
    indra_op_nature = \
        [{'partner_a': i.split('_')[0],
          'partner_b': i.split('_')[1]}
         for i in set(indra_db_interactions.interactions) | set(nature_interactions.interactions)]
    indra_op_nature = pd.DataFrame(indra_op_nature)
    indra_op_nature_uniprot = []
    indra_op_nature_interactions = []

    count = 0
    for r, c in indra_op_nature.iterrows():
        count += 1
        if c[0] in up_hgnc and c[1] in up_hgnc:
            indra_op_nature_uniprot.append(
                {
                    'id_cp_interaction': 'Woolf-' + str(count),
                    'partner_a': up_hgnc[c[0]],
                    'partner_b': up_hgnc[c[1]],
                    'source': 'INDRA'
                }
            )

            indra_op_nature_interactions.append(
                {
                    'id_cp_interaction': 'Woolf-' + str(count),
                    'partner_a': c[0],
                    'partner_b': c[1],
                }
            )

    pd.DataFrame(indra_op_nature_interactions).to_csv(os.path.join(
        HERE, os.pardir, 'output/indra_op_nature_interactions.csv'), sep=",", index=False)

    pd.DataFrame(indra_op_nature_uniprot).to_csv(os.path.join(
        HERE, os.pardir, 'output/indra_op_nature_uniprot.csv'), sep=",", index=False)
